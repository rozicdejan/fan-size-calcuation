import streamlit as st
import pandas as pd
import json, math, io, tempfile
import plotly.io as pio
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib import colors
import plotly.graph_objects as go

# --------------------------
# Helper Functions
# --------------------------
def safe_write_image(fig, path, fmt="png", scale=2):
    try:
        fig.write_image(path, format=fmt, scale=scale)
    except RuntimeError:
        # fallback if no Chrome/Chromium is installed
        with open(path, "w") as f:
            f.write(fig.to_html(full_html=False, include_plotlyjs="cdn"))

def calc_airflow(power_w, delta_t):
    """Required airflow from heat load and allowed ΔT."""
    if delta_t <= 0 or power_w <= 0:
        return 0, 0
    return 3.0 * power_w / delta_t, 1.76 * power_w / delta_t  # (m³/h, CFM)

def calc_surface_area(h, w, d):  # m²
    return 2 * (h*w + h*d + w*d)

def calc_volume(h, w, d):  # m³
    return h * w * d

def calc_dew_point(temp_c, rh):
    """Magnus formula (°C)."""
    a, b = 17.27, 237.7
    gamma = (a * temp_c) / (b + temp_c) + math.log(rh / 100.0)
    return (b * gamma) / (a - gamma)

def calc_heater_power(volume_m3, dew_point, min_temp):
    """Rule of thumb: 10 W per m³ per K above dew point (+2 °C safety)."""
    delta_t = max(0, dew_point - min_temp + 2)
    return 10 * volume_m3 * delta_t

def load_fan_db(file_path="fans.json"):
    try:
        with open(file_path, "r") as f:
            return json.load(f)
    except Exception as e:
        st.error(f"Error loading fan database: {e}")
        return []

def interpolated_flow_at_pressure(fan, static_pressure_pa):
    """Linear interpolate between free-air and 50 Pa; extrapolate beyond 50 Pa."""
    q0 = fan.get("flow_free_m3h", 0.0)
    q50 = fan.get("flow_50pa_m3h", q0)
    p = max(0.0, static_pressure_pa)

    if p <= 50:
        return q0 - (p/50.0)*(q0 - q50)
    # extrapolate linearly past 50 Pa using same slope
    slope = (q50 - q0) / 50.0  # negative
    q = q50 + (p - 50.0) * slope
    return max(q, 0.0)

def select_fans(flow_req_m3h, fans, static_pressure_pa=0):
    """
    Choose a single fan model and count n s.t. n*Q(p) >= required.
    Returns (recommendation_dict, fan_record) or (None, None)
    """
    if not fans or flow_req_m3h <= 0:
        return None, None

    candidates = []
    for fan in fans:
        q = interpolated_flow_at_pressure(fan, static_pressure_pa)
        n = math.ceil(flow_req_m3h / q) if q > 0 else 999999
        total_q = n * q
        total_pwr = n * fan.get("power_w", 0)
        candidates.append({
            "model": fan.get("model", "Unknown"),
            "n": int(n),
            "flow_total": int(round(total_q)),
            "power_total": float(total_pwr),
            "fan_data": fan
        })

    best = sorted(candidates, key=lambda x: (x["flow_total"], x["power_total"]))[0]
    return best, best["fan_data"]

def generate_pq_curves(fan, n_fans=1):
    """
    P–Q curves (all fans & one-fan-failed) from 0..100 Pa using interpolation model.
    Returns pressures, flows_all, flows_fail (or None if n_fans==1)
    """
    pressures = list(range(0, 101, 5))
    flows_all = [interpolated_flow_at_pressure(fan, p)*n_fans for p in pressures]
    flows_fail = [interpolated_flow_at_pressure(fan, p)*(n_fans-1) for p in pressures] if n_fans > 1 else None
    return pressures, flows_all, flows_fail

# --------------------------
# Export Helpers
# --------------------------
def export_excel(data_dict):
    """Styled Excel datasheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Fan Sizing"

    ws.append(["Parameter", "Value"])
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for k, v in data_dict.items():
        ws.append([k, v])

    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = thin
            cell.alignment = Alignment(horizontal="left", vertical="center")

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

def export_pdf(data_dict, redundancy_info=None, airflow_fig=None, redundancy_fig=None, pq_fig=None):
    """PDF report with tables + charts."""
    bio = io.BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("🌀 Electrical Cabinet Fan Sizing Report", styles['Title']))
    elements.append(Spacer(1, 12))

    # Summary table
    table_data = [["Parameter", "Value"]]
    for k, v in data_dict.items():
        table_data.append([k, str(v)])

    table = Table(table_data, colWidths=[200, 250])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#4f81bd")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("ALIGN", (0,0), (-1,-1), "LEFT"),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 10),
        ("BOTTOMPADDING", (0,0), (-1,0), 8),
        ("BACKGROUND", (0,1), (-1,-1), colors.whitesmoke),
        ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 16))

    # Redundancy section
    if redundancy_info:
        elements.append(Paragraph("🛡️ Redundancy Check", styles['Heading2']))
        elements.append(Paragraph("PASS ✅" if redundancy_info["pass"] else "FAIL ⚠️", styles['Normal']))
        elements.append(Spacer(1, 10))
        table2_data = [
            ["Case", "Airflow (m³/h)"],
            ["All fans running", str(int(redundancy_info["all_fans"]))],
            ["1 fan failed", str(int(redundancy_info["one_failed"]))],
            ["Required airflow", str(int(redundancy_info["required"]))],
        ]
        table2 = Table(table2_data, colWidths=[200, 250])
        table2.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#c0504d")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("ALIGN", (0,0), (-1,-1), "LEFT"),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,-1), 10),
            ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
        ]))
        elements.append(table2)
        elements.append(Spacer(1, 16))

    # Charts
    def add_fig(fig, title):
        if not fig:
            return
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            safe_write_image(fig, tmp.name, fmt="png", scale=2)
            elements.append(Paragraph(title, styles['Heading2']))
            elements.append(Image(tmp.name, width=400, height=250))
            elements.append(Spacer(1, 16))

    add_fig(airflow_fig, "📊 Required Airflow vs ΔT")
    add_fig(redundancy_fig, "🛡️ Redundancy Chart")
    add_fig(pq_fig, "📈 Fan P–Q Curve")

    doc.build(elements)
    bio.seek(0)
    return bio

# --------------------------
# Streamlit App
# --------------------------
st.set_page_config(page_title="Cabinet Fan Sizing", layout="wide")
page = st.sidebar.radio("📑 Pages", ["Fan Sizing Calculator", "Fan Database Explorer"])

# --------------------------
# Page 1: Fan Sizing Calculator
# --------------------------
if page == "Fan Sizing Calculator":
    st.title("🌀 Electrical Cabinet Climate Tool (Rittal + Schrack DB)")

    # Power input
    method = st.sidebar.radio("Power input method", ["Sum of devices (W)", "Per device entry", "Upload Excel"])
    total_power = 0.0
    if method == "Sum of devices (W)":
        total_power = st.number_input("Total wattage (W)", min_value=0.0, value=300.0, step=10.0)
    elif method == "Per device entry":
        if "devices" not in st.session_state: st.session_state.devices = [50.0, 100.0, 150.0]
        for i, val in enumerate(st.session_state.devices):
            st.session_state.devices[i] = st.number_input(f"Device {i+1} (W)", min_value=0.0, value=val, step=1.0)
        c1, c2 = st.columns(2)
        with c1:
            if st.button("➕ Add Device"): st.session_state.devices.append(0.0)
        with c2:
            if st.button("➖ Remove Last") and len(st.session_state.devices) > 1: st.session_state.devices.pop()
        total_power = sum(st.session_state.devices)
    else:
        file = st.file_uploader("Upload Excel with a 'Wattage' column", type=["xlsx"])
        if file:
            df = pd.read_excel(file)
            if "Wattage" in df.columns:
                total_power = float(df["Wattage"].sum())
                st.dataframe(df)
            else:
                st.error("Missing 'Wattage' column!")

    # Cabinet environment
    indoor_outdoor = st.radio("Cabinet location", ["Indoor", "Outdoor"])
    solar_power = 0.0
    volume = 0.0
    if indoor_outdoor == "Outdoor":
        st.markdown("### Cabinet Geometry")
        h = st.number_input("Height (m)", value=2.0, step=0.1)
        w = st.number_input("Width (m)", value=1.0, step=0.1)
        d = st.number_input("Depth (m)", value=0.6, step=0.1)
        area = calc_surface_area(h, w, d)
        volume = calc_volume(h, w, d)
        absorptivity = st.select_slider("Paint absorptivity", options=[0.3, 0.5, 0.7, 0.9], value=0.7)
        irradiance = st.number_input("Solar irradiance (W/m²)", value=700.0, step=50.0)
        solar_power = absorptivity * area * irradiance
        st.info(f"Surface area ≈ {area:.2f} m², Volume ≈ {volume:.2f} m³, Solar gain ≈ {solar_power:.0f} W")

    # Temperature & humidity
    st.markdown("### Temperature & Humidity")
    ambient = st.number_input("Ambient temperature (°C)", value=30.0)
    max_inside = st.number_input("Max cabinet temperature (°C)", value=40.0)
    humidity = st.slider("Relative Humidity (%)", 10, 100, 60)

    dew_point = calc_dew_point(ambient, humidity)
    st.write(f"🌡️ Dew point ≈ **{dew_point:.1f} °C**")
    heater_power = calc_heater_power(volume, dew_point, ambient) if volume > 0 else 0.0
    if dew_point >= ambient:
        st.error(f"⚠️ Condensation risk! Suggest heater ≈ {heater_power:.0f} W")
    elif dew_point >= ambient - 2:
        st.warning("⚠️ Dew point close to ambient. Monitor humidity.")
    else:
        st.success("✅ Low condensation risk.")

    delta_t = max_inside - ambient
    total_load = total_power + solar_power
    st.markdown("### Heat Load")
    st.write(f"Devices: **{total_power:.1f} W**, Solar: **{solar_power:.1f} W**, Total: **{total_load:.1f} W**")

    # Static pressure
    static_pressure = st.number_input("Expected static pressure drop (Pa)", value=50, step=10)

    # Fan selection & charts
    airflow_fig = None
    redundancy_fig = None
    pq_fig = None
    redundancy_info = None

    fans = load_fan_db("fans.json")
    if total_load > 0 and delta_t > 0:
        airflow_m3h, airflow_cfm = calc_airflow(total_load, delta_t)
        st.success(f"Required airflow: **{airflow_m3h:.0f} m³/h** (≈ {airflow_cfm:.0f} CFM)")

        margin_pct = st.slider("Safety margin (%)", 0, 100, 30, step=5)
        airflow_with_margin = airflow_m3h * (1 + margin_pct/100.0)
        st.info(f"Recommended airflow with margin: **{airflow_with_margin:.0f} m³/h**")

        rec, fan_detail = select_fans(airflow_with_margin, fans, static_pressure)
        if rec:
            st.subheader("Recommended Fan Configuration")
            st.success(
                f"✅ **{rec['n']} × {rec['model']}**  "
                f"(Total ≈ **{rec['flow_total']} m³/h**, **{rec['power_total']:.0f} W**)"
            )
            if fan_detail:
                extras = []
                for key, label in {
                    "brand": "Brand", "series": "Series", "voltage": "Voltage",
                    "ip_rating": "IP Rating", "size_mm": "Size (mm)",
                    "noise_dba": "Noise (dB(A))", "mtbf_h": "MTBF (h)"
                }.items():
                    if key in fan_detail:
                        extras.append(f"- {label}: **{fan_detail[key]}**")
                if extras: st.markdown("\n".join(extras))

        # Airflow vs ΔT chart
        st.subheader("📊 Airflow Requirement vs ΔT")
        delta_range = list(range(5, 31))
        airflow_curve = [calc_airflow(total_load, dt)[0] for dt in delta_range]
        airflow_fig = go.Figure()
        airflow_fig.add_trace(go.Scatter(x=delta_range, y=airflow_curve, mode="lines+markers", name="Required airflow"))
        airflow_fig.add_trace(go.Scatter(x=[delta_t], y=[airflow_m3h], mode="markers+text",
                                         text=[f"{airflow_m3h:.0f} m³/h"], textposition="top center",
                                         name="Chosen ΔT"))
        airflow_fig.update_layout(xaxis_title="ΔT (°C)", yaxis_title="Airflow (m³/h)",
                                  template="plotly_white")
        st.plotly_chart(airflow_fig, use_container_width=True)

        # Redundancy chart
        if rec and rec["n"] > 1:
            fail_one_flow = (rec["n"] - 1) * (rec["flow_total"] / rec["n"])
            redundancy_info = {
                "all_fans": rec["flow_total"],
                "one_failed": fail_one_flow,
                "required": airflow_with_margin,
                "pass": fail_one_flow >= airflow_with_margin
            }
            redundancy_fig = go.Figure()
            redundancy_fig.add_trace(go.Bar(x=["All fans", "1 failed"],
                                            y=[rec["flow_total"], fail_one_flow],
                                            name="Available airflow"))
            redundancy_fig.add_trace(go.Scatter(x=["All fans", "1 failed"],
                                                y=[airflow_with_margin, airflow_with_margin],
                                                mode="lines", name="Required airflow",
                                                line=dict(dash="dash")))
            redundancy_fig.update_layout(yaxis_title="Airflow (m³/h)", template="plotly_white")
            st.plotly_chart(redundancy_fig, use_container_width=True)
            st.success("✅ Redundancy OK" if redundancy_info["pass"] else "⚠️ Redundancy NOT sufficient")
        elif rec:
            st.info("Redundancy check needs ≥ 2 fans.")

        # P–Q curve with shading & one-failed curve
        if rec and fan_detail:
            st.subheader("📈 Fan P–Q Curve (with redundancy & requirement)")
            pressures, flows_all, flows_fail = generate_pq_curves(fan_detail, rec["n"])

            pq_fig = go.Figure()
            # All fans curve
            pq_fig.add_trace(go.Scatter(x=pressures, y=flows_all, mode="lines+markers",
                                        name=f"All {rec['n']} fans", line=dict(color="blue")))
            # One failed
            if flows_fail:
                pq_fig.add_trace(go.Scatter(x=pressures, y=flows_fail, mode="lines+markers",
                                            name=f"{rec['n']-1} fans (1 failed)",
                                            line=dict(color="orange", dash="dash")))
            # Requirement line
            pq_fig.add_trace(go.Scatter(x=[0, max(pressures)], y=[airflow_with_margin, airflow_with_margin],
                                        mode="lines", name="Required airflow",
                                        line=dict(color="red", dash="dot")))
            # Shading: green above requirement under "all fans"
            pq_fig.add_trace(go.Scatter(
                x=pressures + pressures[::-1],
                y=flows_all + [airflow_with_margin]*len(pressures),
                fill="toself", fillcolor="rgba(0,200,0,0.12)",
                line=dict(color="rgba(0,0,0,0)"), hoverinfo="skip", showlegend=False
            ))
            # Shading: red below requirement
            pq_fig.add_trace(go.Scatter(
                x=pressures + pressures[::-1],
                y=[0]*len(pressures) + [airflow_with_margin]*len(pressures),
                fill="toself", fillcolor="rgba(200,0,0,0.10)",
                line=dict(color="rgba(0,0,0,0)"), hoverinfo="skip", showlegend=False
            ))
            pq_fig.update_layout(xaxis_title="Static Pressure (Pa)", yaxis_title="Airflow (m³/h)",
                                 template="plotly_white")
            st.plotly_chart(pq_fig, use_container_width=True)

        # Export data
        export_data = {
            "Device Dissipation (W)": f"{total_power:.1f}",
            "Solar Gain (W)": f"{solar_power:.1f}",
            "Total Heat Load (W)": f"{total_load:.1f}",
            "Ambient Temp (°C)": f"{ambient:.1f}",
            "Relative Humidity (%)": f"{humidity}",
            "Dew Point (°C)": f"{dew_point:.1f}",
            "Heater Power Suggestion (W)": f"{heater_power:.0f}",
            "Max Cabinet Temp (°C)": f"{max_inside:.1f}",
            "Delta T (°C)": f"{delta_t:.1f}",
            "Static Pressure (Pa)": f"{static_pressure}",
            "Required Airflow (m³/h)": f"{airflow_m3h:.0f}",
            "Required Airflow (CFM)": f"{airflow_cfm:.0f}",
            "Airflow with Margin (m³/h)": f"{airflow_with_margin:.0f}"
        }
        if rec:
            export_data["Fan Recommendation"] = f"{rec['n']} × {rec['model']}"
            export_data["Total Fan Flow (m³/h)"] = f"{rec['flow_total']}"
            export_data["Total Fan Power (W)"] = f"{rec['power_total']:.0f}"
            if fan_detail:
                for k, label in {
                    "brand": "Brand", "series": "Series", "voltage": "Voltage",
                    "ip_rating": "IP Rating", "size_mm": "Size (mm)",
                    "noise_dba": "Noise (dB(A))", "mtbf_h": "MTBF (h)"
                }.items():
                    if k in fan_detail:
                        export_data[label] = str(fan_detail[k])

        # Export buttons
        c1, c2 = st.columns(2)
        with c1:
            st.download_button("📊 Download Excel",
                               export_excel(export_data),
                               file_name="fan_sizing.xlsx")
        with c2:
            st.download_button("📄 Download PDF",
                               export_pdf(export_data, redundancy_info, airflow_fig, redundancy_fig, pq_fig),
                               file_name="fan_sizing.pdf",
                               mime="application/pdf")
    else:
        st.warning("Please input valid wattage and ensure ΔT > 0.")

# --------------------------
# Page 2: Fan Database Explorer
# --------------------------
elif page == "Fan Database Explorer":
    st.title("🔍 Fan Database Explorer")

    fans = load_fan_db("fans.json")
    if not fans:
        st.error("No fans found in fans.json")
    else:
        df = pd.DataFrame(fans)

        st.sidebar.subheader("Filter Options")
        brands = st.sidebar.multiselect("Brand", options=sorted(df["brand"].dropna().unique().tolist()))
        voltages = st.sidebar.multiselect("Voltage", options=sorted(df["voltage"].dropna().unique().tolist()))
        ip_ratings = st.sidebar.multiselect("IP Rating", options=sorted(df["ip_rating"].dropna().unique().tolist()))

        # Range defaults guarded against empty/NaN
        flow_min = int(df["flow_free_m3h"].min()) if len(df) else 0
        flow_max = int(df["flow_free_m3h"].max()) if len(df) else 1000
        pwr_min = int(df["power_w"].min()) if len(df) else 0
        pwr_max = int(df["power_w"].max()) if len(df) else 500

        min_flow, max_flow = st.sidebar.slider(
            "Flow (m³/h, free-air)", flow_min, flow_max, (flow_min, flow_max)
        )
        min_power, max_power = st.sidebar.slider(
            "Power (W)", pwr_min, pwr_max, (pwr_min, pwr_max)
        )

        filtered = df.copy()
        if brands:
            filtered = filtered[filtered["brand"].isin(brands)]
        if voltages:
            filtered = filtered[filtered["voltage"].isin(voltages)]
        if ip_ratings:
            filtered = filtered[filtered["ip_rating"].isin(ip_ratings)]
        filtered = filtered[
            (filtered["flow_free_m3h"] >= min_flow) &
            (filtered["flow_free_m3h"] <= max_flow) &
            (filtered["power_w"] >= min_power) &
            (filtered["power_w"] <= max_power)
        ]

        st.dataframe(filtered)

        csv = filtered.to_csv(index=False).encode("utf-8")
        st.download_button("⬇️ Download filtered fans (CSV)", csv, "filtered_fans.csv", "text/csv")
